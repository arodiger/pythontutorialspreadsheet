import openpyxl                                                         #import library

inv_file = openpyxl.load_workbook("inventory.xlsx")                     #open workbook
product_list = inv_file["Sheet1"]                                       #load worksheet

supplier_dict = {}    
productcount_under_10 = {}                                                  #declare dictionary object


for product_row in range( 2, product_list.max_row + 1 ) :               #iterate thru all the rows of the worksheet
    supplier_name = product_list.cell(product_row, 4).value             #grab value supplier name
    inv_price = product_list.cell(product_row, 3).value                 #grab value inventory price
    inv_count = product_list.cell(product_row, 2).value                 #grab value inventory count
    product_number = product_list.cell(product_row, 1).value            #grab value product number
    running_sum = inv_count * inv_price                                 #compute running total
    
    spreadsheet_inv_price =  product_list.cell(product_row, 5)          #grab the cell object for the working row                 
    spreadsheet_inv_price.value = running_sum                           #update spreadsheet with the running total per inventory item per row

    if (supplier_dict.get(supplier_name) == None) :                     #check for first entry 
        supplier_dict.update({supplier_name : running_sum })            #assign first entry
    else :                                                              #else 
        supplier_dict_value = supplier_dict.get(supplier_name)          #get key's current value
        supplier_dict_value += running_sum                              #compute key's new value with running total
        supplier_dict.update({supplier_name : supplier_dict_value })    #update key's value

    if ( inv_count < 10 ):                                              #any inventory count under 10 needs to be logged
        productcount_under_10.update({int(product_number) : int(inv_count)})      #assign value to dict obj

inv_file.save("temp.xlsx")                                              #create new spreadsheet file with the newly created column

print(supplier_dict)                                                    #printout the total inventory cost per company                                
print(productcount_under_10)                                            #printout dictionary object with totals
print(f"testing git")
print(f"testing git02")
